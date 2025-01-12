import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from utils.database import get_report_data

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
REPORTS_PATH = os.path.join(PROJECT_ROOT, '.reports_cache')


async def create_report(start_date: datetime | None = None, end_date: datetime | None = None) -> str:
    data = await get_report_data(start_date, end_date)
    excel_data = [
        (
            record["user_name"],
            record["fragment_count"],
            record["paid_tokens_spent"],
            record["paid_tokens"],
            record["free_tokens"]
        )
        for record in data
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = ["Пользователь", "Количество отрывков", "Потраченных платных токенов",
               "Платных токенов на счете", "Бесплатных токенов на счете"]
    sheet.append(headers)

    for row in excel_data:
        sheet.append(row)

    header_fill = PatternFill(start_color="A9C6E7", end_color="A9C6E7", fill_type="solid")  # Header fill
    row_fill_odd = PatternFill(start_color="E6F0F7", end_color="E6F0F7", fill_type="solid")  # Odd rows fill
    row_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Even rows fill

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=5), start=2):
        for cell in row:
            if row_index % 2 == 0:
                cell.fill = row_fill_even
            else:
                cell.fill = row_fill_odd

    table = Table(displayName="UserDataTable", ref=f"A1:E{len(data) + 1}")
    sheet.add_table(table)

    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Adjust column widths
    for col_num, col_cells in enumerate(sheet.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_num)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 10
        sheet.column_dimensions[col_letter].width = adjusted_width

    os.makedirs(REPORTS_PATH, exist_ok=True)

    start = "-from-" + start_date.strftime("%d-%m-%Y") if start_date else ''
    end = "-to-" + end_date.strftime("%d-%m-%Y") if end_date else ''
    filename = f"uchibot-report{start}{end}.xlsx"
    workbook.save(os.path.join(REPORTS_PATH, filename))
    return os.path.join(REPORTS_PATH, filename)
